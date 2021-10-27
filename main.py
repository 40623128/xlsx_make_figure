from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np

if __name__ == '__main__':
    #讀取excel
    workbook = load_workbook('座標.xlsx')
    #讀取所有工作表名稱
    sheets = workbook.get_sheet_names()
    #讀取"工作表2"
    booksheet = workbook.get_sheet_by_name("工作表2")
    #booksheet = workbook.get_sheet_by_name(sheets[2])
    x_data =[]
    y_data =[]
    z_data =[]
    #讀取x座標值
    i = 2
    while booksheet.cell(row=i, column=1).value:
        data = booksheet.cell(row=i, column=1).value
        x_data.append(data)
        i = i+1
    #讀取y座標值
    i = 2
    while booksheet.cell(row=i, column=2).value:
        data = booksheet.cell(row=i, column=2).value
        #print(i,' = ',data)
        y_data.append(data)
        i = i+1
    #讀取z座標值
    i = 2
    while booksheet.cell(row=i, column=3).value:
        data = booksheet.cell(row=i, column=3).value
        #print(i,' = ',data)
        z_data.append(data)
        i = i+1
    if len(x_data)==len(y_data)==len(z_data):
        #圖型繪製
        fig = plt.figure()
        ax = fig.add_subplot(projection='3d')
        #marker=點圖示,s=點大小,alpha=透明度
        ax.scatter(x_data, y_data, z_data, marker='o',s=5,alpha = 0.1)

        ax.set_xlabel('X Label')
        ax.set_ylabel('Y Label')
        ax.set_zlabel('Z Label')
        plt.show()
    else:
        print("The amount of data is different")
